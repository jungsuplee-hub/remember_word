"""Administrator endpoints."""
from __future__ import annotations

import asyncio
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Literal
from urllib.parse import quote
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import models
import schemas
from database import get_db
from utils.auth import require_admin
from utils.hanja_lookup import contains_hanja, lookup_meaning

HANJA_JOB_TTL = timedelta(hours=1)


@dataclass
class HanjaMeaningJob:
    """Represents an async job that fills Hanja meanings."""

    id: str
    original_filename: str
    status: Literal["pending", "processing", "completed", "failed"] = "pending"
    created_at: datetime = field(default_factory=datetime.utcnow)
    updated_at: datetime = field(default_factory=datetime.utcnow)
    completed_at: datetime | None = None
    total: int = 0
    processed: int = 0
    filled: int = 0
    existing: int = 0
    missing: int = 0
    message: str | None = None
    download_name: str | None = None
    fallback_name: str | None = None
    download_bytes: bytes | None = None


hanja_jobs: dict[str, HanjaMeaningJob] = {}
hanja_jobs_lock = Lock()


def _cleanup_jobs() -> None:
    """Remove jobs that are older than the TTL to free memory."""

    cutoff = datetime.utcnow() - HANJA_JOB_TTL
    with hanja_jobs_lock:
        expired = [
            job_id
            for job_id, job in hanja_jobs.items()
            if job.completed_at and job.completed_at < cutoff
        ]
        for job_id in expired:
            hanja_jobs.pop(job_id, None)


def _store_job(job: HanjaMeaningJob) -> HanjaMeaningJob:
    with hanja_jobs_lock:
        hanja_jobs[job.id] = job
    return job


def _get_job(task_id: str) -> HanjaMeaningJob:
    with hanja_jobs_lock:
        job = hanja_jobs.get(task_id)
        if job is None:
            raise HTTPException(404, "요청한 작업을 찾을 수 없습니다.")
        return job


def _update_job(task_id: str, **changes) -> HanjaMeaningJob:
    with hanja_jobs_lock:
        job = hanja_jobs.get(task_id)
        if job is None:
            raise HTTPException(404, "요청한 작업을 찾을 수 없습니다.")
        for key, value in changes.items():
            setattr(job, key, value)
        job.updated_at = datetime.utcnow()
        return job


def _serialize_job(job: HanjaMeaningJob) -> dict[str, object]:
    expires_at: datetime | None = None
    if job.completed_at:
        expires_at = job.completed_at + HANJA_JOB_TTL

    return {
        "task_id": job.id,
        "status": job.status,
        "total": job.total,
        "processed": job.processed,
        "filled": job.filled,
        "existing": job.existing,
        "missing": job.missing,
        "message": job.message,
        "created_at": job.created_at.isoformat(),
        "updated_at": job.updated_at.isoformat(),
        "completed_at": job.completed_at.isoformat() if job.completed_at else None,
        "expires_at": expires_at.isoformat() if expires_at else None,
        "download_ready": job.status == "completed" and job.download_bytes is not None,
        "download_name": job.download_name,
    }


def _prepare_download_names(filename: str) -> tuple[str, str]:
    base_name = Path(filename).stem or "hanja"
    safe_base = "".join(
        ch if ch.isalnum() or ch in {"_", "-", " ", "."} else "_" for ch in base_name
    ).strip("_ ")
    if not safe_base:
        safe_base = "hanja"
    download_name = f"{safe_base}_뜻자동입력.xlsx"
    ascii_fallback = "".join(
        ch if ch.isascii() and (ch.isalnum() or ch in {"_", "-", " ", "."}) else "_"
        for ch in safe_base
    ).strip("_ ")
    if not ascii_fallback:
        ascii_fallback = "hanja"
    fallback_name = f"{ascii_fallback}_meaning.xlsx"
    return download_name, fallback_name


def _process_hanja_meaning_job(task_id: str, content: bytes, filename: str) -> None:
    job = _update_job(task_id, status="processing", message=None)

    try:
        workbook = load_workbook(BytesIO(content))
    except InvalidFileException as exc:  # pragma: no cover - openpyxl specific message
        _update_job(task_id, status="failed", message=f"엑셀 파일을 열 수 없습니다: {exc}", completed_at=datetime.utcnow())
        return

    worksheet = workbook.active

    processed_rows: list[tuple[int, str]] = []
    filled = 0
    existing = 0
    missing = 0

    for row in worksheet.iter_rows(min_row=1):
        if not row:
            continue
        term_cell = row[0]
        meaning_cell = row[1] if len(row) > 1 else None

        term_value = term_cell.value
        if term_value is None:
            continue

        term_text = str(term_value).strip()
        if not term_text:
            continue

        if meaning_cell is None:
            meaning_cell = worksheet.cell(row=term_cell.row, column=2)

        existing_value = meaning_cell.value
        if isinstance(existing_value, str):
            existing_text = existing_value.strip()
        else:
            existing_text = str(existing_value).strip() if existing_value is not None else ""

        if existing_text:
            existing += 1
            continue

        processed_rows.append((term_cell.row, term_text))

    total = len(processed_rows)
    _update_job(task_id, total=total, existing=existing)

    for index, (row_number, term_text) in enumerate(processed_rows, start=1):
        translated = lookup_meaning(term_text)
        if translated:
            worksheet.cell(row=row_number, column=2, value=translated)
            filled += 1
        else:
            if contains_hanja(term_text):
                missing += 1
        _update_job(
            task_id,
            processed=index,
            filled=filled,
            missing=missing,
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    download_name, fallback_name = _prepare_download_names(filename)

    _update_job(
        task_id,
        status="completed",
        completed_at=datetime.utcnow(),
        download_bytes=output.getvalue(),
        download_name=download_name,
        fallback_name=fallback_name,
        message="총 {}건 중 {}건에 뜻을 입력했습니다.".format(total, filled)
        if total
        else "채워 넣을 항목이 없습니다.",
    )


async def _run_hanja_meaning_job(task_id: str, content: bytes, filename: str) -> None:
    try:
        await asyncio.to_thread(_process_hanja_meaning_job, task_id, content, filename)
    except Exception as exc:  # pragma: no cover - defensive
        _update_job(
            task_id,
            status="failed",
            message=f"알 수 없는 오류가 발생했습니다: {exc}",
            completed_at=datetime.utcnow(),
        )

router = APIRouter()


@router.get("/dashboard", response_model=list[schemas.AdminAccountStats])
def dashboard(
    _: models.Profile = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[schemas.AdminAccountStats]:
    profiles = db.query(models.Profile).order_by(models.Profile.id).all()

    folder_counts = dict(
        db.query(models.Folder.profile_id, func.count(models.Folder.id))
        .group_by(models.Folder.profile_id)
        .all()
    )
    group_counts = dict(
        db.query(models.Group.profile_id, func.count(models.Group.id))
        .group_by(models.Group.profile_id)
        .all()
    )
    word_counts = dict(
        db.query(models.Group.profile_id, func.count(models.Word.id))
        .select_from(models.Word)
        .join(models.Group, models.Group.id == models.Word.group_id)
        .group_by(models.Group.profile_id)
        .all()
    )
    quiz_counts = dict(
        db.query(models.QuizSession.profile_id, func.count(models.QuizSession.id))
        .filter(models.QuizSession.is_completed.is_(True))
        .group_by(models.QuizSession.profile_id)
        .all()
    )

    stats: list[schemas.AdminAccountStats] = []
    for profile in profiles:
        pid = profile.id
        stats.append(
            schemas.AdminAccountStats(
                profile_id=pid,
                username=profile.username,
                name=profile.name,
                email=profile.email,
                folder_count=folder_counts.get(pid, 0),
                group_count=group_counts.get(pid, 0),
                word_count=word_counts.get(pid, 0),
                quiz_count=quiz_counts.get(pid, 0),
                login_count=profile.login_count or 0,
                last_login_at=profile.last_login_at,
            )
        )

    return stats


@router.post(
    "/utilities/hanja-meanings",
    status_code=202,
    response_model=schemas.HanjaMeaningJobCreated,
)
async def populate_hanja_meanings(
    _: models.Profile = Depends(require_admin),
    file: UploadFile = File(...),
) -> schemas.HanjaMeaningJobCreated:
    """Start an asynchronous job that populates Hanja meanings in an Excel sheet."""

    _cleanup_jobs()

    if not file.filename:
        raise HTTPException(400, "업로드할 엑셀 파일을 선택하세요.")

    filename = file.filename or ""
    lower_name = filename.lower()
    if not lower_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(400, "엑셀(.xlsx) 파일만 지원합니다.")

    content = await file.read()

    task_id = uuid4().hex
    job = HanjaMeaningJob(id=task_id, original_filename=filename)
    _store_job(job)

    asyncio.create_task(_run_hanja_meaning_job(task_id, content, filename))

    return schemas.HanjaMeaningJobCreated(
        task_id=task_id,
        status_url=f"/admin/utilities/hanja-meanings/{task_id}",
        download_url=f"/admin/utilities/hanja-meanings/{task_id}/download",
    )


@router.get(
    "/utilities/hanja-meanings/{task_id}",
    response_model=schemas.HanjaMeaningJobStatus,
)
async def get_hanja_meaning_status(
    task_id: str,
    _: models.Profile = Depends(require_admin),
) -> schemas.HanjaMeaningJobStatus:
    _cleanup_jobs()
    job = _get_job(task_id)
    return schemas.HanjaMeaningJobStatus(**_serialize_job(job))


@router.get("/utilities/hanja-meanings/{task_id}/download", response_class=StreamingResponse)
async def download_hanja_meaning_result(
    task_id: str,
    _: models.Profile = Depends(require_admin),
) -> StreamingResponse:
    job = _get_job(task_id)
    if job.status != "completed" or job.download_bytes is None:
        raise HTTPException(409, "작업이 아직 완료되지 않았습니다.")

    download_name, fallback_name = job.download_name, job.fallback_name
    if not download_name or not fallback_name:
        download_name, fallback_name = _prepare_download_names(job.original_filename)

    quoted_name = quote(download_name)

    response = StreamingResponse(
        BytesIO(job.download_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers[
        "Content-Disposition"
    ] = f"attachment; filename=\"{fallback_name}\"; filename*=UTF-8''{quoted_name}"

    response.headers["X-Meaning-Processed"] = str(job.processed)
    response.headers["X-Meaning-Filled"] = str(job.filled)
    response.headers["X-Meaning-Existing"] = str(job.existing)
    response.headers["X-Meaning-Missing"] = str(job.missing)

    return response
